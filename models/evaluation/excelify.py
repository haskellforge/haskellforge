"""
Convert a .json file with the following format for each line:
{"input": "...", "gt": "...", "prediction": "..."}

to an Excel file with the following format for each line:
input | gt | prediction

Example usage from root directory:
python3 models/evaluation/excelify.py -f ./models/evaluation/data/codegpt_base-test-humaneval.json ./models/evaluation/data/codegpt_finetuned-test-humaneval.json ./models/evaluation/data/unixcoder_base-test-humaneval.json ./models/evaluation/data/unixcoder_finetuned-test-humaneval.json -o ./models/evaluation/output
"""
import json
import os
from argparse import ArgumentParser

import openpyxl
from fuzzywuzzy import fuzz


def main():
    """
    Generates Excel file(s).

    Raises:
        ValueError: files to convert do not exist.
    """
    parser = ArgumentParser()
    parser.add_argument('-f', '--file', required=True, nargs="+")
    parser.add_argument('-o', '--output-folder', required=True)
    args = parser.parse_args()

    if not os.path.exists(args.output_folder):
        os.makedirs(args.output_folder)

    for file in args.file:
        if not os.path.exists(file):
            raise ValueError(f"File {file} does not exist.")

    for file in args.file:
        output_file_path = os.path.join(
            args.output_folder, f"{os.path.basename(file).split('.')[0]}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active

        overlap_file = None
        overlaps_folder = os.path.join(os.path.dirname(
            os.path.abspath(__file__)), "overlap-check")
        if "unixcoder" in file:
            overlap_file = os.path.join(
                overlaps_folder, "overlapping-unixcoder_finetuned-test-humaneval.json")
        elif "codegpt" in file:
            overlap_file = os.path.join(
                overlaps_folder, "overlapping-codegpt_finetuned-test-humaneval.json")

        overlap_data = None
        if overlap_file is not None:
            with open(overlap_file, 'r') as of:
                overlap_data = json.load(of)

        # Get the data from the .json file and convert it into Excel format
        with open(file) as f:
            lines = [json.loads(line) for line in f]

            # Write the header
            headers = ["Input", "Ground Truth", "Prediction",
                       "EM", "ES", "OL"]  # EM = Exact Match, ES = Exact Similarity, OL = Overlap

            categories = {
                "if/then/else": ["all", "if", "then", "else"],
                "generators": ["complete", "body"],
                "guards (= |)": ["complete", "body"],
                "functions": ["complete", "parameter(s)", "argument(s)", "body"],
                "lists": ["++", ":", "!!", "list comprehension", "range"],
                "logical operators": ["all", "&&", "||", "==", ">", "<", ">=", "<=", "/=", "not"],
                "arithmetic operators": ["all", "+", "-", "*", "/", "^", "mod"],
                "case expressions": ["complete", "parameter(s)", "argument(s)"],
                "other": ["variable name", "wrong type", "wrong value"],
            }
            headers.append("")
            for category in categories:
                headers.append(category + "\n\n"
                               + "\n".join(categories[category]))
            headers.append("")
            headers.append("other comments")
            ws.append(headers)

            # Make header titles bold and bigger font and wrap text and align center horizontally and vertically
            for cell in ws["1:1"]:
                cell.font = openpyxl.styles.Font(bold=True, size=14)
                cell.alignment = openpyxl.styles.Alignment(
                    wrap_text=True, horizontal="center", vertical="center")
                cell.fill = openpyxl.styles.fills.PatternFill(
                    patternType='solid', fgColor="DDDDDD")

            # Make width of columns G to O as wide as the longest categories[category] string
            for i in range(8, 17):
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    i)].width = 20

            # Make G1 background white
            ws["G1"].fill = openpyxl.styles.fills.PatternFill(
                patternType='solid', fgColor="FFFFFF")
            ws["Q1"].fill = openpyxl.styles.fills.PatternFill(
                patternType='solid', fgColor="FFFFFF")

            # Light green and width of 100
            ws["Q1"].fill = openpyxl.styles.fills.PatternFill(
                patternType='solid', fgColor="DDFFDD")
            ws.column_dimensions["R"].width = 100

            # Freeze the first row while scrolling
            ws.freeze_panes = "A2"

            # Write the data
            for line in lines:

                # Calculate EM and ES
                em = "True" if line["prediction"].strip(
                ) == line["gt"].strip() else "False"
                es = fuzz.ratio(line["prediction"].strip(),
                                line["gt"].strip()) / 100
                ol = "False"
                if overlap_data is not None:
                    for overlap_line in overlap_data:
                        if overlap_line["input"].strip() == line["input"].strip() and overlap_line["prediction"].strip() == line["prediction"].strip():
                            ol = "True"
                            break

                ws.append([
                    readable_json_value(line["input"]),
                    readable_json_value(line["gt"]),
                    readable_json_value(line["prediction"]),
                    em, es, ol])

            # Adjust the column widths and wrap text
            ws.column_dimensions["A"].width = 125
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 50
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

            # Center the text in the cells of columns EM and ES
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal="center")

            # If the value in the fourth column is "True", color the cell green, otherwise red, and allow custom color formatting
            green = openpyxl.styles.colors.Color(rgb='00FF00')
            red = openpyxl.styles.colors.Color(rgb='FF0000')
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    if cell.value == "True":
                        cell.fill = openpyxl.styles.fills.PatternFill(
                            patternType='solid', fgColor=green)
                    else:
                        cell.fill = openpyxl.styles.fills.PatternFill(
                            patternType='solid', fgColor=red)
            for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
                for cell in row:
                    if cell.value == "True":
                        cell.fill = openpyxl.styles.fills.PatternFill(
                            patternType='solid', fgColor=green)
                    else:
                        cell.fill = openpyxl.styles.fills.PatternFill(
                            patternType='solid', fgColor=red)

            # Save the Excel file
            wb.save(output_file_path)


def readable_json_value(input: str) -> str:
    """
    Improve readability of the JSON input by removing EOL tokens and <s> and </s> tokens.

    Args:
        input (str): raw JSON input.

    Returns:
        str: formatted JSON input.
    """
    return input.replace("<EOL>", "\n").replace("<s>", "").replace("</s>", "")


if __name__ == "__main__":
    main()
